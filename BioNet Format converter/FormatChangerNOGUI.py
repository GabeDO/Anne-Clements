# -*- coding: utf-8 -*-
"""
Created on Tue Jun  1 10:40:41 2021

@author: z5035086
"""
    
import os
import openpyxl


file = os.getcwd()+'\Template.xlsx'
Input = os.getcwd()+'\Maroota Quarry 2020.xlsx'


SiteName = "ACA_M_"
Date = '18/11/2020 00:00:00'
NameCol = 3    #Col with plant names


row = 4
wbInput = openpyxl.load_workbook(filename = Input)
wsInput = wbInput.active
row_count = wsInput.max_row

wbTemp = openpyxl.load_workbook(filename=file)
wsTemp = wbTemp.active

for N in range(int(wsInput.max_column)):
    print(N)
    row = 4
    new_row = 69
    for R in range(row_count):       
        InputRow = list(wsInput[int(NameCol)+R])
        if  InputRow[N].value == "<1":
            new_row = [None,SiteName+str(list(wsInput[1])[N].value),str(Date),None ,None ,None ,None ,None ,str(InputRow[NameCol-1].value).rstrip(),None ,None ,None ,None ,None ,None ,0.01,None ,None ,4,None ,None ,None ,None ]     
            print(new_row[1], new_row[8])
            for col, entry in enumerate(new_row, start=1):
                wsTemp.cell(row=row, column=col, value=entry)
            row = row + 1 
            #wbTemp.save(filename='Data\\'+str(SiteName+str(list(wsInput[1])[N].value))+'.xlsx')                        
        
        elif  isinstance(InputRow[N].value, str) != True and InputRow[N].value != None:            
            new_row = [None,SiteName+str(list(wsInput[1])[N].value),str(Date),None ,None ,None ,None ,None ,str(InputRow[NameCol-1].value).rstrip(),None ,None ,None ,None ,None ,None ,InputRow[N].value,None ,None ,4,None ,None ,None ,None ]     
            print(new_row[1], new_row[8])
            for col, entry in enumerate(new_row, start=1):
                wsTemp.cell(row=row, column=col, value=entry)
            row = row + 1
            
    if new_row != 69:
        wbTemp.save(filename='Data\\'+str(SiteName+str(list(wsInput[1])[N].value))+'.xlsx')
    
    wsTemp = wbTemp.active


    
   
