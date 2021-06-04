# -*- coding: utf-8 -*-
"""
Created on Tue Jun  1 10:40:41 2021

@author: z5035086
"""
import tkinter as tk

root= tk.Tk()

canvas1 = tk.Canvas(root, width = 400, height = 300)
canvas1.pack()

label2 = tk.Label(root, text="This program will convert Anne's feild data \n into the format required to upload to BioNet. \n The input data must be called 'Input.xlxs'.")
label2.config(font=('helvetica', 10))
canvas1.create_window(200, 60, window=label2)

label2 = tk.Label(root, text='Date of collection \n (dd/mm/yyyy hh:mm:ss)')
label2.config(font=('helvetica', 10))
canvas1.create_window(200, 140, window=label2)
DateEntry1 = tk.Entry (root) 
canvas1.create_window(200, 170, window=DateEntry1)


label2 = tk.Label(root, text='Coloum number of plant names')
label2.config(font=('helvetica', 10))
canvas1.create_window(200, 230, window=label2)
NameColSource = tk.Entry (root) 
canvas1.create_window(200, 250, window=NameColSource)
    
import os

import openpyxl

def RunCode ():  
    Date = DateEntry1.get()
    NameCol = NameColSource.get() 
    label1 = tk.Label(root, text= "Run")

    file = os.getcwd()+'\Template.xlsx'
    Input = os.getcwd()+'\Input.xlsx'
    
    
    #NumberOfSites = input("Number of sites:")
    #Date = input("Date of collection:")
    row = 4
    
    wbInput = openpyxl.load_workbook(filename = Input)
    wsInput = wbInput.active
    row_count = wsInput.max_row
    
    wbTemp = openpyxl.load_workbook(filename=file)
    wsTemp = wbTemp.active
    
    for R in range(row_count):
        InputRow = list(wsInput[int(NameCol)+R])
        print(R)
        for N in range(int(wsInput.max_column)):
            
            if  InputRow[N].value == "<1":
                new_row = [None,list(wsInput[1])[N].value,str(Date),None ,None ,None ,None ,None ,str(InputRow[1].value).rstrip(),None ,None ,None ,None ,None ,None ,9999,None ,None ,4,None ,None ,None ,None ]     
                for col, entry in enumerate(new_row, start=1):
                    wsTemp.cell(row=row, column=col, value=entry)
                row = row + 1
            
            elif  isinstance(InputRow[N].value, str) != True and InputRow[N].value != None:            
                new_row = [None,list(wsInput[1])[N].value,str(Date),None ,None ,None ,None ,None ,str(InputRow[1].value).rstrip(),None ,None ,None ,None ,None ,None ,InputRow[N].value,None ,None ,4,None ,None ,None ,None ]     
                for col, entry in enumerate(new_row, start=1):
                    wsTemp.cell(row=row, column=col, value=entry)
                row = row + 1
                
        
    wbTemp.save(filename='Output.xlsx')
   
button1 = tk.Button(text='Run', command=RunCode)
canvas1.create_window(200, 280, window=button1)

root.mainloop()